"""
API路由定义（V2.0 升级）
保留V1全部功能 + 新增客户发现、搜索任务管理、断点续跑等API
"""
import json
import datetime
import os
import asyncio
from typing import Optional, Dict, Set
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from sqlalchemy.orm import Session
from app.database import get_db, Customer, SearchTask
from app.services.excel_importer import parse_excel, import_customers
from app.services.website_scraper import scrape_website
from app.services.email_extractor import extract_emails_from_text
from app.services.keyword_analyzer import analyze_keywords
from app.services.deepseek_analyzer import analyze_company, generate_summary
from app.services.scoring_engine import calculate_scores
from app.services.search_task_service import run_search_task, request_stop, get_stop_flag, get_paused_tasks, resume_paused_task
from app.services.keyword_expander import expand_keywords
from app.services.keyword_analyzer import analyze_keywords
from app.services.scoring_engine import calculate_scores

router = APIRouter(prefix="/api", tags=["customers"])

# ──── 全局分析状态控制 ────
_analyzing_set: Set[int] = set()
_running_tasks: Set[int] = set()


def _get_emails_list(customer: Customer) -> list:
    if not customer.emails:
        return []
    try:
        return json.loads(customer.emails)
    except (json.JSONDecodeError, TypeError):
        return [e.strip() for e in customer.emails.split(",") if e.strip()]


# ═══════════════════════════════════════════
# V1 保留功能：客户管理
# ═══════════════════════════════════════════

@router.get("/customers")
def list_customers(
    search: Optional[str] = Query(None, description="搜索关键词（支持公司名/网址/邮箱）"),
    country: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    status: Optional[str] = Query(None, description="跟进状态筛选: 待联系/已发邮件/已回复/无效线索/成单"),
    sort_by_score: Optional[str] = Query("desc"),
    page: int = Query(1, ge=1, description="页码，从1开始"),
    page_size: int = Query(50, ge=10, le=200, description="每页条数（10-200）"),
    db: Session = Depends(get_db),
):
    query = db.query(Customer)
    if search:
        query = query.filter(
            Customer.company_name.ilike(f"%{search}%")
            | Customer.website.ilike(f"%{search}%")
            | Customer.emails.ilike(f"%{search}%")
        )
    if country:
        query = query.filter(Customer.country == country)
    if priority:
        query = query.filter(Customer.priority == priority.upper())
    if status:
        query = query.filter(Customer.status == status)
    if sort_by_score == "asc":
        query = query.order_by(Customer.total_score.asc().nullslast())
    else:
        query = query.order_by(Customer.total_score.desc().nullslast())

    total_count = query.count()
    offset_val = (page - 1) * page_size
    customers = query.offset(offset_val).limit(page_size).all()

    # 统计数据合并返回
    stats = {
        "total": db.query(Customer).count(),
        "analyzed": db.query(Customer).filter(Customer.analyzed_at.isnot(None)).count(),
        "grade_a": db.query(Customer).filter(Customer.priority == "A").count(),
        "google": db.query(Customer).filter(Customer.discovery_source == "Google").count(),
    }

    all_countries = db.query(Customer.country).distinct().filter(
        Customer.country.isnot(None), Customer.country != ""
    ).order_by(Customer.country).all()
    country_list = [c[0] for c in all_countries if c[0]]

    result = []
    for c in customers:
        emails = _get_emails_list(c)
        result.append({
            "id": c.id,
            "company_name": c.company_name,
            "website": c.website or "",
            "country": c.country or "",
            "email_count": len(emails),
            "total_score": c.total_score,
            "priority": c.priority or "-",
            "ai_summary": c.ai_summary or "",
            "discovery_source": c.discovery_source or "",
            "discovery_keyword": c.discovery_keyword or "",
            "is_analyzing": c.id in _analyzing_set,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "analyzed_at": c.analyzed_at.isoformat() if c.analyzed_at else None,
            "status": c.status or "待联系",
            "follow_up_date": c.follow_up_date.isoformat() if c.follow_up_date else None,
            "scrape_status": c.scrape_status,
            "ai_status": c.ai_status,
            "fail_reason": c.fail_reason,
            "star_rating": c.star_rating or 0,
        })

    return {
        "customers": result,
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total_count + page_size - 1) // page_size),
        "countries": country_list,
        "stats": stats,
    }


@router.get("/customers/{customer_id}")
def get_customer_detail(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    emails = _get_emails_list(customer)

    positive_kw = {}
    if customer.positive_keywords:
        try:
            positive_kw = json.loads(customer.positive_keywords)
        except (json.JSONDecodeError, TypeError):
            pass

    negative_kw = {}
    if customer.negative_keywords:
        try:
            negative_kw = json.loads(customer.negative_keywords)
        except (json.JSONDecodeError, TypeError):
            pass

    ai_raw = {}
    if customer.ai_raw_json:
        try:
            ai_raw = json.loads(customer.ai_raw_json)
        except (json.JSONDecodeError, TypeError):
            pass

    return {
        "id": customer.id,
        "company_name": customer.company_name,
        "website": customer.website or "",
        "country": customer.country or "",
        "emails": emails,
        "website_text": customer.website_text or "",
        "positive_keywords": positive_kw,
        "negative_keywords": negative_kw,
        "industry_score": customer.industry_score,
        "project_score": customer.project_score,
        "company_type_score": customer.company_type_score,
        "country_score": customer.country_score,
        "contact_score": customer.contact_score,
        "total_score": customer.total_score,
        "priority": customer.priority or "-",
        "company_type": customer.company_type or "",
        "ai_summary": customer.ai_summary or "",
        "sales_hook": customer.sales_hook or "",
        "target_position": customer.target_position or "",
        "identified_projects": customer.identified_projects or "",
        "discovery_source": customer.discovery_source or "",
        "discovery_keyword": customer.discovery_keyword or "",
        "first_found_at": customer.first_found_at.isoformat() if customer.first_found_at else None,
        "ai_raw": ai_raw,
        "is_analyzing": customer.id in _analyzing_set,
        "created_at": customer.created_at.isoformat() if customer.created_at else None,
        "analyzed_at": customer.analyzed_at.isoformat() if customer.analyzed_at else None,
        # V2.2 跟进状态
        "status": customer.status or "待联系",
        "follow_up_date": customer.follow_up_date.isoformat() if customer.follow_up_date else None,
        "notes": customer.notes or "",
        # V2.2 抓取/分析状态
        "scrape_status": customer.scrape_status,
        "ai_status": customer.ai_status,
        "fail_reason": customer.fail_reason,
        # V2.2 客户评级
        "star_rating": customer.star_rating or 0,
    }


@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 格式的Excel文件")
    os.makedirs("app/uploads", exist_ok=True)
    upload_path = f"app/uploads/{file.filename}"
    content = await file.read()
    with open(upload_path, "wb") as f:
        f.write(content)
    try:
        customers_data = parse_excel(upload_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析Excel失败: {str(e)}")
    if not customers_data:
        raise HTTPException(status_code=400, detail="Excel中没有找到有效数据")
    count = import_customers(customers_data)
    return {"message": f"成功导入 {count} 个客户", "total_in_file": len(customers_data), "imported": count, "skipped": len(customers_data) - count}


@router.post("/analyze/{customer_id}")
async def analyze_single(customer_id: int, db: Session = Depends(get_db)):
    if customer_id in _analyzing_set:
        raise HTTPException(status_code=400, detail="该客户正在分析中")
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    if not customer.website:
        raise HTTPException(status_code=400, detail="该客户没有官网地址")
    _analyzing_set.add(customer_id)
    try:
        website_text = await scrape_website(customer.website)
        if website_text:
            customer.scrape_status = "success"
            customer.website_text = website_text
            emails = extract_emails_from_text(website_text)
            email_list = list(set(emails))
            customer.emails = json.dumps(email_list, ensure_ascii=False)
            pos_hits, neg_hits = analyze_keywords(website_text)
            customer.positive_keywords = json.dumps(pos_hits, ensure_ascii=False)
            customer.negative_keywords = json.dumps(neg_hits, ensure_ascii=False)
            ai_result = await analyze_company(website_text)
            if ai_result:
                customer.ai_status = "success"
                customer.ai_raw_json = json.dumps(ai_result, ensure_ascii=False)
                customer.company_type = ai_result.get("company_type", "")
                customer.sales_hook = ai_result.get("sales_hook", "")
                customer.target_position = ai_result.get("target_position", "")
                customer.identified_projects = ai_result.get("identified_projects", "")
                customer_info = {"country": customer.country or "", "company_name": customer.company_name}
                customer.ai_summary = generate_summary(ai_result, customer_info)
            else:
                customer.ai_status = "failed"
                customer.fail_reason = "AI分析失败（API可能超时）"
            scores = calculate_scores(
                website_text=website_text, positive_keywords=pos_hits,
                company_type=customer.company_type, country=customer.country, emails=email_list,
            )
            customer.industry_score = scores["industry_score"]
            customer.project_score = scores["project_score"]
            customer.company_type_score = scores["company_type_score"]
            customer.country_score = scores["country_score"]
            customer.contact_score = scores["contact_score"]
            customer.total_score = scores["total_score"]
            customer.priority = scores["priority"]
        customer.analyzed_at = datetime.datetime.utcnow()
        db.commit()
        return {"message": "分析完成", "customer_id": customer.id}
    except Exception as e:
        db.rollback()
        raise e
    finally:
        _analyzing_set.discard(customer_id)


@router.post("/analyze-all")
async def analyze_all(db: Session = Depends(get_db)):
    customers = db.query(Customer).filter(Customer.analyzed_at.is_(None)).all()
    if not customers:
        return {"message": "没有待分析的客户", "analyzed_count": 0}
    analyzed_count = 0
    for customer in customers:
        if not customer.website:
            continue
        if customer.id in _analyzing_set:
            continue
        try:
            await analyze_single(customer.id, db)
            analyzed_count += 1
        except Exception as e:
            print(f"分析 {customer.company_name} 失败: {str(e)}")
            continue
    return {"message": "批量分析完成", "analyzed_count": analyzed_count}


@router.post("/stop-analysis")
async def stop_analysis():
    request_stop()
    return {"message": "已发送停止信号，正在等待当前任务完成"}


@router.get("/analysis-status")
def get_analysis_status():
    return {
        "is_analyzing": len(_analyzing_set) > 0,
        "analyzing_ids": list(_analyzing_set),
    }


@router.get("/export-excel")
def export_excel(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    from fastapi.responses import StreamingResponse
    customers = db.query(Customer).order_by(Customer.total_score.desc().nullslast()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户分析结果"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ["公司名称", "国家", "官网", "邮箱数量", "总分", "优先级", "公司类型", "发现来源", "发现关键词", "AI摘要", "开发切入点", "推荐联系职位", "分析时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
    for row_idx, c in enumerate(customers, 2):
        emails = _get_emails_list(c)
        row_data = [
            c.company_name, c.country or "", c.website or "", len(emails),
            c.total_score or 0, c.priority or "-", c.company_type or "",
            c.discovery_source or "", c.discovery_keyword or "",
            c.ai_summary or "", c.sales_hook or "", c.target_position or "",
            c.analyzed_at.strftime("%Y-%m-%d %H:%M") if c.analyzed_at else "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border; cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate([25, 15, 35, 10, 8, 8, 18, 12, 20, 35, 30, 20, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=customers_export.xlsx"})


@router.delete("/customers/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    db.delete(customer); db.commit()
    return {"message": "删除成功"}


@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    total = db.query(Customer).count()
    analyzed = db.query(Customer).filter(Customer.analyzed_at.isnot(None)).count()
    pending = total - analyzed
    a = db.query(Customer).filter(Customer.priority == "A").count()
    b = db.query(Customer).filter(Customer.priority == "B").count()
    c = db.query(Customer).filter(Customer.priority == "C").count()
    d = db.query(Customer).filter(Customer.priority == "D").count()
    google_count = db.query(Customer).filter(Customer.discovery_source == "Google").count()
    manual_count = db.query(Customer).filter(Customer.discovery_source.is_(None)).count()
    return {
        "total": total, "analyzed": analyzed, "pending": pending,
        "priority_distribution": {"A": a, "B": b, "C": c, "D": d},
        "discovery_stats": {"google": google_count, "manual_import": manual_count},
    }


# ═══════════════════════════════════════════
# V2.0 新增：客户发现 & 搜索任务管理
# ═══════════════════════════════════════════

@router.post("/discovery/expand-keywords")
async def api_expand_keywords(
    keyword: str = Query(..., description="需要扩展的基础关键词"),
    country: str = Query("", description="目标国家（可选，传入后AI会用该国家的本地语言扩展关键词）"),
):
    """AI扩展关键词（支持多语言：传入 country 会自动使用该国语言）"""
    expanded = await expand_keywords(keyword, country=country)
    if not expanded:
        expanded = [keyword]
    return {"original_keyword": keyword, "expanded_keywords": expanded, "country": country}


@router.post("/discovery/search-task")
async def create_search_task(
    country: str = Query(..., description="搜索国家"),
    keyword: str = Query(..., description="搜索关键词"),
    depth: int = Query(50, description="搜索深度"),
    db: Session = Depends(get_db),
):
    """创建新的搜索发现任务"""
    task = SearchTask(
        country=country,
        keyword=keyword,
        search_depth=depth,
        status="Pending",
        created_at=datetime.datetime.utcnow(),
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    # 异步启动任务
    loop = asyncio.get_event_loop()
    loop.create_task(_run_task_wrapper(task.id))

    return {"message": "搜索任务已创建", "task_id": task.id}


async def _run_task_wrapper(task_id: int):
    """异步包装器，确保任务异常不影响主进程"""
    try:
        await run_search_task(task_id)
    except Exception as e:
        print(f"搜索任务 {task_id} 异常退出: {str(e)[:200]}")


@router.get("/discovery/tasks")
def list_search_tasks(db: Session = Depends(get_db)):
    """获取所有搜索任务列表"""
    tasks = db.query(SearchTask).order_by(SearchTask.created_at.desc()).all()
    result = []
    for t in tasks:
        expanded = []
        if t.expanded_keywords:
            try:
                expanded = json.loads(t.expanded_keywords)
            except (json.JSONDecodeError, TypeError):
                pass
        result.append({
            "id": t.id,
            "country": t.country,
            "keyword": t.keyword,
            "expanded_keywords": expanded,
            "search_depth": t.search_depth,
            "status": t.status,
            "found_websites": t.found_websites or 0,
            "analyzed_companies": t.analyzed_companies or 0,
            "new_companies": t.new_companies or 0,
            "current_keyword_index": t.current_keyword_index or 0,
            "error_message": t.error_message,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "finished_at": t.finished_at.isoformat() if t.finished_at else None,
            "task_log": t.task_log or "",
        })

    # 标记当前活跃任务（优先 Running，其次 Pending）
    active_task_id = None
    for t in tasks:
        if t.status == "Running":
            active_task_id = t.id
            break
    if active_task_id is None:
        for t in tasks:
            if t.status == "Pending":
                active_task_id = t.id
                break

    return {"tasks": result, "total": len(result), "active_task_id": active_task_id}


@router.post("/discovery/tasks/{task_id}/pause")
def pause_task(task_id: int, db: Session = Depends(get_db)):
    """暂停搜索任务"""
    request_stop()
    return {"message": "正在停止搜索任务"}


@router.post("/discovery/tasks/{task_id}/resume")
def resume_task(task_id: int, db: Session = Depends(get_db)):
    """恢复暂停的搜索任务（断点续跑）"""
    success = resume_paused_task(task_id)
    if not success:
        raise HTTPException(status_code=400, detail="任务不存在或无法恢复")
    # 重新启动
    loop = asyncio.get_event_loop()
    loop.create_task(_run_task_wrapper(task_id))
    return {"message": "任务已恢复", "task_id": task_id}


@router.delete("/discovery/tasks/{task_id}")
def delete_task(task_id: int, db: Session = Depends(get_db)):
    """删除搜索任务"""
    task = db.query(SearchTask).filter(SearchTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    db.delete(task)
    db.commit()
    return {"message": "删除成功"}


@router.get("/discovery/paused-tasks")
def list_paused_tasks(db: Session = Depends(get_db)):
    """获取所有暂停的任务（用于断点续跑）"""
    tasks = get_paused_tasks(db)
    result = []
    for t in tasks:
        result.append({
            "id": t.id,
            "country": t.country,
            "keyword": t.keyword,
            "status": t.status,
            "current_keyword_index": t.current_keyword_index or 0,
        })
    return {"paused_tasks": result}


@router.get("/discovery/discovered-customers")
def list_discovered_customers(
    search: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None, description="按发现关键词筛选"),
    sort_by_score: Optional[str] = Query("desc"),
    db: Session = Depends(get_db),
):
    """获取通过Google发现的公司列表"""
    query = db.query(Customer).filter(Customer.discovery_source == "Google")
    if search:
        query = query.filter(Customer.company_name.ilike(f"%{search}%"))
    if country:
        query = query.filter(Customer.country == country)
    if priority:
        query = query.filter(Customer.priority == priority.upper())
    if keyword:
        query = query.filter(Customer.discovery_keyword.ilike(f"%{keyword}%"))
    if sort_by_score == "asc":
        query = query.order_by(Customer.total_score.asc().nullslast())
    else:
        query = query.order_by(Customer.total_score.desc().nullslast())

    customers = query.all()
    all_countries = db.query(Customer.country).distinct().filter(
        Customer.country.isnot(None), Customer.country != "",
        Customer.discovery_source == "Google",
    ).order_by(Customer.country).all()
    country_list = [c[0] for c in all_countries if c[0]]

    all_keywords = db.query(Customer.discovery_keyword).distinct().filter(
        Customer.discovery_keyword.isnot(None), Customer.discovery_keyword != "",
        Customer.discovery_source == "Google",
    ).all()
    keyword_list = list(set(k[0] for k in all_keywords if k[0]))

    result = []
    for c in customers:
        emails = _get_emails_list(c)
        result.append({
            "id": c.id,
            "company_name": c.company_name,
            "website": c.website or "",
            "country": c.country or "",
            "email_count": len(emails),
            "total_score": c.total_score,
            "priority": c.priority or "-",
            "discovery_keyword": c.discovery_keyword or "",
            "ai_summary": c.ai_summary or "",
            "created_at": c.created_at.isoformat() if c.created_at else None,
        })

    return {"customers": result, "total": len(result), "countries": country_list, "keywords": keyword_list}


# ═══════════════════════════════════════════
# V2.2 新增：客户跟进状态管理
# ═══════════════════════════════════════════

@router.post("/customers/{customer_id}/follow-up")
def update_follow_up(
    customer_id: int,
    status: str = Query(..., description="跟进状态: 待联系/已发邮件/已回复/无效线索/成单"),
    follow_up_date: Optional[str] = Query(None, description="下次跟进日期 (YYYY-MM-DD)"),
    notes: Optional[str] = Query("", description="跟进备注"),
    star_rating: Optional[int] = Query(None, description="客户评级 (0=未评级, 1-5星)"),
    db: Session = Depends(get_db),
):
    """更新客户跟进状态和评级"""
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    valid_statuses = ["待联系", "已发邮件", "已回复", "无效线索", "成单"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"无效状态，可选: {'/'.join(valid_statuses)}")

    customer.status = status
    if follow_up_date:
        try:
            customer.follow_up_date = datetime.datetime.strptime(follow_up_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="日期格式错误，请使用 YYYY-MM-DD")
    if notes:
        customer.notes = notes
    if star_rating is not None:
        if star_rating < 0 or star_rating > 5:
            raise HTTPException(status_code=400, detail="评级范围 0-5")
        customer.star_rating = star_rating
    db.commit()

    return {"message": "跟进状态已更新", "customer_id": customer.id, "status": customer.status, "star_rating": customer.star_rating}


# ═══════════════════════════════════════════
# V2.2 新增：重新抓取 / 重新分析（局部重试）
# ═══════════════════════════════════════════

@router.post("/customers/{customer_id}/re-scrape")
async def rescrape_customer(customer_id: int, db: Session = Depends(get_db)):
    """重新抓取客户官网（保留已有AI分析结果，仅刷新抓取数据）"""
    if customer_id in _analyzing_set:
        raise HTTPException(status_code=400, detail="该客户正在分析中")
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    if not customer.website:
        raise HTTPException(status_code=400, detail="该客户没有官网地址")

    _analyzing_set.add(customer_id)
    try:
        domain = customer.website.replace("https://", "").replace("http://", "").split("/")[0]
        website_url = f"https://{domain}"
        website_text = await scrape_website(website_url)

        if website_text:
            customer.scrape_status = "success"
            customer.website_text = website_text
            # 重新提取邮箱
            emails = extract_emails_from_text(website_text)
            email_list = list(set(emails))
            customer.emails = json.dumps(email_list, ensure_ascii=False)
            # 重新关键词分析
            pos_hits, neg_hits = analyze_keywords(website_text)
            customer.positive_keywords = json.dumps(pos_hits, ensure_ascii=False)
            customer.negative_keywords = json.dumps(neg_hits, ensure_ascii=False)
            # 重新评分
            scores = calculate_scores(
                website_text=website_text, positive_keywords=pos_hits,
                company_type=customer.company_type, country=customer.country, emails=email_list,
            )
            customer.industry_score = scores["industry_score"]
            customer.project_score = scores["project_score"]
            customer.company_type_score = scores["company_type_score"]
            customer.country_score = scores["country_score"]
            customer.contact_score = scores["contact_score"]
            customer.total_score = scores["total_score"]
            customer.priority = scores["priority"]
        else:
            customer.scrape_status = "failed"
            customer.fail_reason = "抓取失败（网站可能无法访问）"

        customer.analyzed_at = datetime.datetime.utcnow()
        db.commit()
        return {"message": "重新抓取完成", "customer_id": customer.id, "scrape_status": customer.scrape_status}
    except Exception as e:
        db.rollback()
        customer.scrape_status = "failed"
        customer.fail_reason = str(e)[:200]
        db.commit()
        raise HTTPException(status_code=500, detail=f"重新抓取失败: {str(e)[:200]}")
    finally:
        _analyzing_set.discard(customer_id)


@router.post("/customers/{customer_id}/re-analyze")
async def reanalyze_customer(customer_id: int, db: Session = Depends(get_db)):
    """重新AI分析客户（仅重新调用DeepSeek，不重新抓取）"""
    if customer_id in _analyzing_set:
        raise HTTPException(status_code=400, detail="该客户正在分析中")
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    if not customer.website_text:
        raise HTTPException(status_code=400, detail="没有官网内容，请先抓取官网")

    _analyzing_set.add(customer_id)
    try:
        ai_result = await analyze_company(customer.website_text)
        if ai_result:
            customer.ai_status = "success"
            customer.ai_raw_json = json.dumps(ai_result, ensure_ascii=False)
            customer.company_type = ai_result.get("company_type", "")
            customer.sales_hook = ai_result.get("sales_hook", "")
            customer.target_position = ai_result.get("target_position", "")
            customer.identified_projects = ai_result.get("identified_projects", "")
            customer_info = {"country": customer.country or "", "company_name": customer.company_name}
            customer.ai_summary = generate_summary(ai_result, customer_info)
            # 重新评分（公司类型可能变了）
            pos_hits = {}
            if customer.positive_keywords:
                try:
                    pos_hits = json.loads(customer.positive_keywords)
                except (json.JSONDecodeError, TypeError):
                    pass
            emails = _get_emails_list(customer)
            scores = calculate_scores(
                website_text=customer.website_text, positive_keywords=pos_hits,
                company_type=customer.company_type, country=customer.country, emails=emails,
            )
            customer.industry_score = scores["industry_score"]
            customer.project_score = scores["project_score"]
            customer.company_type_score = scores["company_type_score"]
            customer.country_score = scores["country_score"]
            customer.contact_score = scores["contact_score"]
            customer.total_score = scores["total_score"]
            customer.priority = scores["priority"]
        else:
            customer.ai_status = "failed"
            customer.fail_reason = "AI分析失败（API可能超时）"

        customer.analyzed_at = datetime.datetime.utcnow()
        db.commit()
        return {"message": "重新分析完成", "customer_id": customer.id, "ai_status": customer.ai_status}
    except Exception as e:
        db.rollback()
        customer.ai_status = "failed"
        customer.fail_reason = str(e)[:200]
        db.commit()
        raise HTTPException(status_code=500, detail=f"重新分析失败: {str(e)[:200]}")
    finally:
        _analyzing_set.discard(customer_id)


# ═══════════════════════════════════════════
# V2.5 新增：相似客户扩展（种子客户）
# ═══════════════════════════════════════════

@router.post("/discovery/similar-companies")
async def api_find_similar_companies(
    company_url: str = Query(..., description="目标公司网址"),
    target_country: str = Query(..., description="目标国家"),
    top_n: int = Query(50, ge=10, le=100, description="返回结果数量"),
):
    """基于公司网址寻找相似客户"""
    from app.services.similar_company_finder import find_similar_companies

    result = await find_similar_companies(company_url, target_country, top_n=top_n)
    return result
