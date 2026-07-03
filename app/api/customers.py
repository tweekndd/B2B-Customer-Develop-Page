"""
客户管理 API 路由
包含客户CRUD、分析、导入导出、跟进状态、重新抓取/分析
从 routes.py 拆分（V2.8 重构）
"""
import json
import datetime
import os
from typing import Optional, Set

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, case as sql_case

from app.database import get_db, Customer
from app.services.excel_importer import parse_excel, import_customers
from app.services.website_scraper import scrape_website
from app.services.email_extractor import extract_emails_from_text
from app.services.keyword_analyzer import analyze_keywords
from app.services.glm_analyzer import analyze_company, generate_summary
from app.services.scoring_engine import calculate_scores
from app.services.search_task_service import request_stop
from app.auth import check_ai_analysis_permission

router = APIRouter(tags=["customers"])

# ──── 全局分析状态控制 ────
_analyzing_set: Set[int] = set()
_running_tasks: Set[int] = set()  # 预留，当前未使用


def _get_emails_list(customer: Customer) -> list:
    """安全解析客户邮箱 JSON 字段"""
    if not customer.emails:
        return []
    try:
        return json.loads(customer.emails)
    except (json.JSONDecodeError, TypeError):
        return [e.strip() for e in customer.emails.split(",") if e.strip()]


# ═══════════════════════════════════════════
# 客户列表 & 详情
# ═══════════════════════════════════════════

@router.get("/customers")
def list_customers(
    search: Optional[str] = Query(None, description="搜索关键词（支持公司名/网址/邮箱）"),
    country: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    status: Optional[str] = Query(None, description="跟进状态筛选: 待联系/已发邮件/已回复/无效线索/成单"),
    sort_by_score: Optional[str] = Query("desc"),
    page: int = Query(1, ge=1, description="页码，从1开始"),
    page_size: int = Query(50, ge=10, le=200, description="每页条数（10-200）"),
    db: Session = Depends(get_db),
):
    query = db.query(Customer)
    if search:
        query = query.filter(
            Customer.company_name.ilike(f"%{search}%")
            | Customer.website.ilike(f"%{search}%")
            | Customer.emails.ilike(f"%{search}%")
        )
    if country:
        query = query.filter(Customer.country == country)
    if priority:
        query = query.filter(Customer.priority == priority.upper())
    if status:
        query = query.filter(Customer.status == status)
    if sort_by_score == "asc":
        query = query.order_by(Customer.total_score.asc().nullslast())
    else:
        query = query.order_by(Customer.total_score.desc().nullslast())

    total_count = query.count()
    offset_val = (page - 1) * page_size
    customers = query.offset(offset_val).limit(page_size).all()

    # 单次聚合查询替代 5 次独立 COUNT
    agg = db.query(
        func.count().label("total"),
        func.sum(sql_case((Customer.analyzed_at.isnot(None), 1), else_=0)).label("analyzed"),
        func.sum(sql_case((Customer.priority == "A", 1), else_=0)).label("grade_a"),
        func.sum(sql_case((Customer.discovery_source == "Google", 1), else_=0)).label("google"),
    ).first()
    stats = {
        "total": agg.total,
        "analyzed": agg.analyzed or 0,
        "grade_a": agg.grade_a or 0,
        "google": agg.google or 0,
    }

    all_countries = db.query(Customer.country).distinct().filter(
        Customer.country.isnot(None), Customer.country != ""
    ).order_by(Customer.country).all()
    country_list = [c[0] for c in all_countries if c[0]]

    result = []
    for c in customers:
        emails = _get_emails_list(c)
        result.append({
            "id": c.id,
            "company_name": c.company_name,
            "website": c.website or "",
            "country": c.country or "",
            "email_count": len(emails),
            "total_score": c.total_score,
            "priority": c.priority or "-",
            "ai_summary": c.ai_summary or "",
            "discovery_source": c.discovery_source or "",
            "discovery_keyword": c.discovery_keyword or "",
            "is_analyzing": c.id in _analyzing_set,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "analyzed_at": c.analyzed_at.isoformat() if c.analyzed_at else None,
            "status": c.status or "待联系",
            "follow_up_date": c.follow_up_date.isoformat() if c.follow_up_date else None,
            "scrape_status": c.scrape_status,
            "ai_status": c.ai_status,
            "fail_reason": c.fail_reason,
            "star_rating": c.star_rating or 0,
        })

    return {
        "customers": result,
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total_count + page_size - 1) // page_size),
        "countries": country_list,
        "stats": stats,
    }


@router.get("/customers/{customer_id}")
def get_customer_detail(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    emails = _get_emails_list(customer)

    positive_kw = {}
    if customer.positive_keywords:
        try:
            positive_kw = json.loads(customer.positive_keywords)
        except (json.JSONDecodeError, TypeError):
            pass

    negative_kw = {}
    if customer.negative_keywords:
        try:
            negative_kw = json.loads(customer.negative_keywords)
        except (json.JSONDecodeError, TypeError):
            pass

    ai_raw = {}
    if customer.ai_raw_json:
        try:
            ai_raw = json.loads(customer.ai_raw_json)
        except (json.JSONDecodeError, TypeError):
            pass

    return {
        "id": customer.id,
        "company_name": customer.company_name,
        "website": customer.website or "",
        "country": customer.country or "",
        "emails": emails,
        "website_text": customer.website_text or "",
        "positive_keywords": positive_kw,
        "negative_keywords": negative_kw,
        "industry_score": customer.industry_score,
        "project_score": customer.project_score,
        "company_type_score": customer.company_type_score,
        "country_score": customer.country_score,
        "contact_score": customer.contact_score,
        "total_score": customer.total_score,
        "priority": customer.priority or "-",
        "company_type": customer.company_type or "",
        "ai_summary": customer.ai_summary or "",
        "sales_hook": customer.sales_hook or "",
        "target_position": customer.target_position or "",
        "identified_projects": customer.identified_projects or "",
        "discovery_source": customer.discovery_source or "",
        "discovery_keyword": customer.discovery_keyword or "",
        "first_found_at": customer.first_found_at.isoformat() if customer.first_found_at else None,
        "ai_raw": ai_raw,
        "is_analyzing": customer.id in _analyzing_set,
        "created_at": customer.created_at.isoformat() if customer.created_at else None,
        "analyzed_at": customer.analyzed_at.isoformat() if customer.analyzed_at else None,
        # V2.2 跟进状态
        "status": customer.status or "待联系",
        "follow_up_date": customer.follow_up_date.isoformat() if customer.follow_up_date else None,
        "notes": customer.notes or "",
        # V2.2 抓取/分析状态
        "scrape_status": customer.scrape_status,
        "ai_status": customer.ai_status,
        "fail_reason": customer.fail_reason,
        # V2.2 客户评级
        "star_rating": customer.star_rating or 0,
    }


# ═══════════════════════════════════════════
# Excel 导入
# ═══════════════════════════════════════════

@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 格式的Excel文件")
    os.makedirs("app/uploads", exist_ok=True)
    upload_path = f"app/uploads/{file.filename}"
    content = await file.read()
    with open(upload_path, "wb") as f:
        f.write(content)
    try:
        customers_data = parse_excel(upload_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析Excel失败: {str(e)}")
    if not customers_data:
        raise HTTPException(status_code=400, detail="Excel中没有找到有效数据")
    count = import_customers(customers_data)
    return {"message": f"成功导入 {count} 个客户", "total_in_file": len(customers_data), "imported": count, "skipped": len(customers_data) - count}


# ═══════════════════════════════════════════
# 客户分析（单个 & 批量）
# ═══════════════════════════════════════════

@router.post("/analyze/{customer_id}")
async def analyze_single(
    customer_id: int,
    db: Session = Depends(get_db),
    user=Depends(check_ai_analysis_permission),
):
    if customer_id in _analyzing_set:
        raise HTTPException(status_code=400, detail="该客户正在分析中")
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    if not customer.website:
        raise HTTPException(status_code=400, detail="该客户没有官网地址")
    _analyzing_set.add(customer_id)
    try:
        website_text = await scrape_website(customer.website)
        if website_text:
            customer.scrape_status = "success"
            customer.website_text = website_text
            emails = extract_emails_from_text(website_text)
            email_list = list(set(emails))
            customer.emails = json.dumps(email_list, ensure_ascii=False)
            pos_hits, neg_hits = analyze_keywords(website_text)
            customer.positive_keywords = json.dumps(pos_hits, ensure_ascii=False)
            customer.negative_keywords = json.dumps(neg_hits, ensure_ascii=False)
            ai_result = await analyze_company(website_text)
            if ai_result:
                customer.ai_status = "success"
                customer.ai_raw_json = json.dumps(ai_result, ensure_ascii=False)
                customer.company_type = ai_result.get("company_type", "")
                customer.sales_hook = ai_result.get("sales_hook", "")
                customer.target_position = ai_result.get("target_position", "")
                customer.identified_projects = ai_result.get("identified_projects", "")
                # AI 提取城市（如果有且客户尚无 city）
                ai_city = ai_result.get("address_city", "")
                if ai_city and not customer.city:
                    customer.city = ai_city.strip()
                customer_info = {"country": customer.country or "", "company_name": customer.company_name}
                customer.ai_summary = generate_summary(ai_result, customer_info)
            else:
                customer.ai_status = "failed"
                customer.fail_reason = "AI分析失败（API可能超时）"
            scores = calculate_scores(
                website_text=website_text, positive_keywords=pos_hits,
                company_type=customer.company_type, country=customer.country, emails=email_list,
            )
            customer.industry_score = scores["industry_score"]
            customer.project_score = scores["project_score"]
            customer.company_type_score = scores["company_type_score"]
            customer.country_score = scores["country_score"]
            customer.contact_score = scores["contact_score"]
            customer.total_score = scores["total_score"]
            customer.priority = scores["priority"]
        customer.analyzed_at = datetime.datetime.utcnow()
        db.commit()
        return {"message": "分析完成", "customer_id": customer.id}
    except Exception as e:
        db.rollback()
        raise e
    finally:
        _analyzing_set.discard(customer_id)


@router.post("/analyze-all")
async def analyze_all(
    db: Session = Depends(get_db),
    user=Depends(check_ai_analysis_permission),
):
    customers = db.query(Customer).filter(Customer.analyzed_at.is_(None)).all()
    if not customers:
        return {"message": "没有待分析的客户", "analyzed_count": 0}
    analyzed_count = 0
    for customer in customers:
        if not customer.website:
            continue
        if customer.id in _analyzing_set:
            continue
        try:
            await analyze_single(customer.id, db)
            analyzed_count += 1
        except Exception as e:
            print(f"分析 {customer.company_name} 失败: {str(e)}")
            continue
    return {"message": "批量分析完成", "analyzed_count": analyzed_count}


@router.post("/stop-analysis")
async def stop_analysis():
    request_stop()
    return {"message": "已发送停止信号，正在等待当前任务完成"}


@router.get("/analysis-status")
def get_analysis_status():
    return {
        "is_analyzing": len(_analyzing_set) > 0,
        "analyzing_ids": list(_analyzing_set),
    }


# ═══════════════════════════════════════════
# Excel 导出
# ═══════════════════════════════════════════

@router.get("/export-excel")
def export_excel(db: Session = Depends(get_db)):
    """导出客户列表为 Excel，按以下规则生成列：
       A-Country | B-Company Name | C-二次开发 | D-邮箱 | E-电话 | F-备注 | G-Website | H-领英"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    from fastapi.responses import StreamingResponse

    customers = db.query(Customer).order_by(Customer.total_score.desc().nullslast()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户列表"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 列头（严格按文档规则）
    headers = [
        "Country",          # A列
        "Company Name",     # B列
        "公司概况",          # C列（可空）
        "邮箱",              # D列
        "电话",              # E列（可空）
        "备注",              # F列（可空）
        "Website",          # G列
        "领英",              # H列（可空）
        "跟进状态",          # I列
        "AI评分",            # J列
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, c in enumerate(customers, 2):
        # 邮箱列表转为逗号分隔的文本
        emails = _get_emails_list(c)
        email_text = ", ".join(emails) if emails else ""

        row_data = [
            c.country or "",                          # A: Country
            c.company_name or "",                     # B: Company Name
            c.ai_summary or "",                       # C: 公司概况（可空）
            email_text,                                # D: 邮箱（多个用逗号分隔）
            "",                                        # E: 电话（暂无数据，可空）
            c.notes or "",                             # F: 备注（可空）
            c.website or "",                           # G: Website
            "",                                        # H: 领英（暂无数据，可空）
            c.status or "待联系",                       # I: 跟进状态
            c.total_score if c.total_score is not None else "",  # J: AI评分
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = cell_alignment

    # 列宽（适配各列内容）
    col_widths = [15, 30, 40, 40, 15, 25, 35, 30, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customers_export.xlsx"},
    )


# ═══════════════════════════════════════════
# 客户删除 & 批量删除 & 统计
# ═══════════════════════════════════════════

@router.delete("/customers/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    db.delete(customer); db.commit()
    return {"message": "删除成功"}


@router.post("/customers/batch-delete")
def batch_delete_customers(
    ids: str = Query(..., description="要删除的客户ID列表，JSON数组字符串"),
    db: Session = Depends(get_db),
):
    """批量删除客户（支持全选/多选后统一删除）"""
    try:
        customer_ids = json.loads(ids)
        if not isinstance(customer_ids, list) or not customer_ids:
            raise HTTPException(status_code=400, detail="ids 应为非空JSON数组")
    except (json.JSONDecodeError, ValueError):
        raise HTTPException(status_code=400, detail="ids 参数格式错误")

    deleted = 0
    for cid in customer_ids:
        customer = db.query(Customer).filter(Customer.id == cid).first()
        if customer:
            db.delete(customer)
            deleted += 1
    db.commit()
    return {"message": f"成功删除 {deleted} 个客户", "deleted": deleted}


@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    agg = db.query(
        func.count().label("total"),
        func.sum(sql_case((Customer.analyzed_at.isnot(None), 1), else_=0)).label("analyzed"),
        func.sum(sql_case((Customer.priority == "A", 1), else_=0)).label("a"),
        func.sum(sql_case((Customer.priority == "B", 1), else_=0)).label("b"),
        func.sum(sql_case((Customer.priority == "C", 1), else_=0)).label("c"),
        func.sum(sql_case((Customer.priority == "D", 1), else_=0)).label("d"),
        func.sum(sql_case((Customer.discovery_source == "Google", 1), else_=0)).label("google_count"),
        func.sum(sql_case(
            (Customer.discovery_source.is_(None), 1),
            (Customer.discovery_source == "", 1),
            else_=0,
        )).label("manual_count"),
    ).first()
    return {
        "total": agg.total,
        "analyzed": agg.analyzed or 0,
        "pending": (agg.total or 0) - (agg.analyzed or 0),
        "priority_distribution": {
            "A": agg.a or 0, "B": agg.b or 0, "C": agg.c or 0, "D": agg.d or 0,
        },
        "discovery_stats": {
            "google": agg.google_count or 0,
            "manual_import": agg.manual_count or 0,
        },
    }


# ═══════════════════════════════════════════
# V3.1.2：保存 Hunter 查到的邮箱到客户记录
# ═══════════════════════════════════════════

@router.post("/customers/{customer_id}/add-emails")
def add_customer_emails(
    customer_id: int,
    emails: str = Query(..., description="要添加的邮箱列表，JSON 数组字符串"),
    set_status: Optional[str] = Query(None, description="同时设置跟进状态"),
    db: Session = Depends(get_db),
):
    """将 Hunter 查到的邮箱保存到客户记录，可选同时更新跟进状态"""
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    try:
        new_emails = json.loads(emails)
        if not isinstance(new_emails, list):
            raise ValueError
    except (json.JSONDecodeError, ValueError):
        raise HTTPException(status_code=400, detail="emails 参数应为 JSON 数组字符串")

    # 合并去重
    existing = _get_emails_list(customer)
    merged = list(dict.fromkeys(existing + new_emails))  # 去重保序
    customer.emails = json.dumps(merged, ensure_ascii=False)
    customer.updated_at = datetime.datetime.utcnow()

    # 可选更新跟进状态
    if set_status:
        valid_statuses = ["待联系", "已发邮件", "已回复", "无效线索", "成单"]
        if set_status in valid_statuses:
            customer.status = set_status

    db.commit()
    return {
        "message": f"已添加 {len(new_emails)} 个邮箱，共 {len(merged)} 个",
        "customer_id": customer.id,
        "email_count": len(merged),
        "added": len(new_emails),
        "status": customer.status,
    }


# ═══════════════════════════════════════════
# V2.2：客户跟进状态管理
# ═══════════════════════════════════════════

@router.post("/customers/{customer_id}/follow-up")
def update_follow_up(
    customer_id: int,
    status: str = Query(..., description="跟进状态: 待联系/已发邮件/已回复/无效线索/成单"),
    follow_up_date: Optional[str] = Query(None, description="下次跟进日期 (YYYY-MM-DD)"),
    notes: Optional[str] = Query("", description="跟进备注"),
    star_rating: Optional[int] = Query(None, description="客户评级 (0=未评级, 1-5星)"),
    db: Session = Depends(get_db),
):
    """更新客户跟进状态和评级"""
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    valid_statuses = ["待联系", "已发邮件", "已回复", "无效线索", "成单"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"无效状态，可选: {'/'.join(valid_statuses)}")

    customer.status = status
    if follow_up_date:
        try:
            customer.follow_up_date = datetime.datetime.strptime(follow_up_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="日期格式错误，请使用 YYYY-MM-DD")
    if notes:
        customer.notes = notes
    if star_rating is not None:
        if star_rating < 0 or star_rating > 5:
            raise HTTPException(status_code=400, detail="评级范围 0-5")
        customer.star_rating = star_rating
    db.commit()

    return {"message": "跟进状态已更新", "customer_id": customer.id, "status": customer.status, "star_rating": customer.star_rating}


# ═══════════════════════════════════════════
# V2.2：重新抓取 / 重新分析（局部重试）
# ═══════════════════════════════════════════

@router.post("/customers/{customer_id}/re-scrape")
async def rescrape_customer(customer_id: int, db: Session = Depends(get_db)):
    """重新抓取客户官网（保留已有AI分析结果，仅刷新抓取数据）"""
    if customer_id in _analyzing_set:
        raise HTTPException(status_code=400, detail="该客户正在分析中")
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    if not customer.website:
        raise HTTPException(status_code=400, detail="该客户没有官网地址")

    _analyzing_set.add(customer_id)
    try:
        domain = customer.website.replace("https://", "").replace("http://", "").split("/")[0]
        website_url = f"https://{domain}"
        website_text = await scrape_website(website_url)

        if website_text:
            customer.scrape_status = "success"
            customer.website_text = website_text
            # 重新提取邮箱
            emails = extract_emails_from_text(website_text)
            email_list = list(set(emails))
            customer.emails = json.dumps(email_list, ensure_ascii=False)
            # 重新关键词分析
            pos_hits, neg_hits = analyze_keywords(website_text)
            customer.positive_keywords = json.dumps(pos_hits, ensure_ascii=False)
            customer.negative_keywords = json.dumps(neg_hits, ensure_ascii=False)
            # 重新评分
            scores = calculate_scores(
                website_text=website_text, positive_keywords=pos_hits,
                company_type=customer.company_type, country=customer.country, emails=email_list,
            )
            customer.industry_score = scores["industry_score"]
            customer.project_score = scores["project_score"]
            customer.company_type_score = scores["company_type_score"]
            customer.country_score = scores["country_score"]
            customer.contact_score = scores["contact_score"]
            customer.total_score = scores["total_score"]
            customer.priority = scores["priority"]
        else:
            customer.scrape_status = "failed"
            customer.fail_reason = "抓取失败（网站可能无法访问）"

        customer.analyzed_at = datetime.datetime.utcnow()
        db.commit()
        return {"message": "重新抓取完成", "customer_id": customer.id, "scrape_status": customer.scrape_status}
    except Exception as e:
        db.rollback()
        customer.scrape_status = "failed"
        customer.fail_reason = str(e)[:200]
        db.commit()
        raise HTTPException(status_code=500, detail=f"重新抓取失败: {str(e)[:200]}")
    finally:
        _analyzing_set.discard(customer_id)


@router.post("/customers/{customer_id}/re-analyze")
async def reanalyze_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    user=Depends(check_ai_analysis_permission),
):
    """重新AI分析客户（仅重新调用DeepSeek，不重新抓取）"""
    if customer_id in _analyzing_set:
        raise HTTPException(status_code=400, detail="该客户正在分析中")
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    if not customer.website_text:
        raise HTTPException(status_code=400, detail="没有官网内容，请先抓取官网")

    _analyzing_set.add(customer_id)
    try:
        ai_result = await analyze_company(customer.website_text)
        if ai_result:
            customer.ai_status = "success"
            customer.ai_raw_json = json.dumps(ai_result, ensure_ascii=False)
            customer.company_type = ai_result.get("company_type", "")
            customer.sales_hook = ai_result.get("sales_hook", "")
            customer.target_position = ai_result.get("target_position", "")
            customer.identified_projects = ai_result.get("identified_projects", "")
            # AI 提取城市（如果有且客户尚无 city）
            ai_city = ai_result.get("address_city", "")
            if ai_city and not customer.city:
                customer.city = ai_city.strip()
            customer_info = {"country": customer.country or "", "company_name": customer.company_name}
            customer.ai_summary = generate_summary(ai_result, customer_info)
            # 重新评分（公司类型可能变了）
            pos_hits = {}
            if customer.positive_keywords:
                try:
                    pos_hits = json.loads(customer.positive_keywords)
                except (json.JSONDecodeError, TypeError):
                    pass
            emails = _get_emails_list(customer)
            scores = calculate_scores(
                website_text=customer.website_text, positive_keywords=pos_hits,
                company_type=customer.company_type, country=customer.country, emails=emails,
            )
            customer.industry_score = scores["industry_score"]
            customer.project_score = scores["project_score"]
            customer.company_type_score = scores["company_type_score"]
            customer.country_score = scores["country_score"]
            customer.contact_score = scores["contact_score"]
            customer.total_score = scores["total_score"]
            customer.priority = scores["priority"]
        else:
            customer.ai_status = "failed"
            customer.fail_reason = "AI分析失败（API可能超时）"

        customer.analyzed_at = datetime.datetime.utcnow()
        db.commit()
        return {"message": "重新分析完成", "customer_id": customer.id, "ai_status": customer.ai_status}
    except Exception as e:
        db.rollback()
        customer.ai_status = "failed"
        customer.fail_reason = str(e)[:200]
        db.commit()
        raise HTTPException(status_code=500, detail=f"重新分析失败: {str(e)[:200]}")
    finally:
        _analyzing_set.discard(customer_id)
